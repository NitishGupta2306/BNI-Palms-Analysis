"""Utilities for exporting analysis results to various formats."""

from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from src.domain.models.analysis_result import AnalysisMatrix, MatrixType
from src.domain.models.member import Member
from src.domain.models.tyfcb import TYFCB
from src.domain.services.tyfcb_service import TYFCBService, TYFCBSummary
from src.domain.exceptions.domain_exceptions import ExportError
from src.infrastructure.data.file_handlers.excel_handler import ExcelHandler
from src.shared.constants.app_constants import MatrixHeaders, CombinationValues


class ExportService:
    """Service for exporting analysis results to various formats."""
    
    def __init__(self):
        self.excel_handler = ExcelHandler()
    
    def export_referral_matrix(self, matrix: AnalysisMatrix, file_path: Path) -> None:
        """
        Export referral matrix to Excel file.
        
        Args:
            matrix: Referral analysis matrix
            file_path: Output file path
        """
        try:
            workbook = self.excel_handler.create_styled_workbook()
            worksheet = workbook.active
            worksheet.title = "Referral Matrix"
            
            members = matrix.get_all_members()
            
            # Write matrix data
            self._write_matrix_to_worksheet(worksheet, matrix, members)
            
            # Add summary columns for referrals
            self._add_referral_summary_columns(worksheet, matrix, members)
            
            # Add summary rows for referrals
            self._add_referral_summary_rows(worksheet, matrix, members)
            
            # Apply styling
            data_range = {'max_row': len(members) + 4, 'max_col': len(members) + 4}
            self.excel_handler.apply_matrix_styling(worksheet, data_range)
            self.excel_handler.auto_adjust_column_widths(worksheet)
            
            # Save workbook
            self.excel_handler.save_workbook(workbook, file_path)
            
        except Exception as e:
            raise ExportError(f"Failed to export referral matrix: {str(e)}")
    
    def export_oto_matrix(self, matrix: AnalysisMatrix, file_path: Path) -> None:
        """
        Export one-to-one matrix to Excel file.
        
        Args:
            matrix: One-to-one analysis matrix
            file_path: Output file path
        """
        try:
            workbook = self.excel_handler.create_styled_workbook()
            worksheet = workbook.active
            worksheet.title = "One to One Matrix"
            
            members = matrix.get_all_members()
            
            # Write matrix data
            self._write_matrix_to_worksheet(worksheet, matrix, members)
            
            # Add summary columns for OTO
            self._add_oto_summary_columns(worksheet, matrix, members)
            
            # Apply styling
            data_range = {'max_row': len(members) + 2, 'max_col': len(members) + 3}
            self.excel_handler.apply_matrix_styling(worksheet, data_range)
            self.excel_handler.auto_adjust_column_widths(worksheet)
            
            # Save workbook
            self.excel_handler.save_workbook(workbook, file_path)
            
        except Exception as e:
            raise ExportError(f"Failed to export OTO matrix: {str(e)}")
    
    def export_combination_matrix(self, matrix: AnalysisMatrix, file_path: Path) -> None:
        """
        Export combination matrix to Excel file.
        
        Args:
            matrix: Combination analysis matrix
            file_path: Output file path
        """
        try:
            workbook = self.excel_handler.create_styled_workbook()
            worksheet = workbook.active
            worksheet.title = "Combination Matrix"
            
            members = matrix.get_all_members()
            
            # Write matrix data
            self._write_matrix_to_worksheet(worksheet, matrix, members)
            
            # Add summary columns for combination
            self._add_combination_summary_columns(worksheet, matrix, members)
            
            # Apply styling
            data_range = {'max_row': len(members) + 2, 'max_col': len(members) + 6}
            self.excel_handler.apply_matrix_styling(worksheet, data_range)
            self.excel_handler.auto_adjust_column_widths(worksheet)
            
            # Save workbook
            self.excel_handler.save_workbook(workbook, file_path)
            
        except Exception as e:
            raise ExportError(f"Failed to export combination matrix: {str(e)}")
    
    def _write_matrix_to_worksheet(self, worksheet, matrix: AnalysisMatrix, 
                                 members: List[Member]) -> None:
        """Write the matrix data to a worksheet."""
        try:
            # Write headers
            worksheet.cell(row=1, column=1, value=MatrixHeaders.GIVER_RECEIVER).font = Font(bold=True)
            
            for col, member in enumerate(members, start=2):
                worksheet.cell(row=1, column=col, value=member.full_name).font = Font(bold=True)
            
            # Write matrix data
            for row, giver in enumerate(members, start=2):
                worksheet.cell(row=row, column=1, value=giver.full_name).font = Font(bold=True)
                
                for col, receiver in enumerate(members, start=2):
                    value = matrix.get_cell_value(giver, receiver)
                    cell = worksheet.cell(row=row, column=col, value=value)
                    
                    # Highlight zero values if this is not a combination matrix
                    if value == 0 and matrix.matrix_type != MatrixType.COMBINATION:
                        zero_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell.fill = zero_fill
            
        except Exception as e:
            raise ExportError(f"Error writing matrix to worksheet: {str(e)}")
    
    def _add_referral_summary_columns(self, worksheet, matrix: AnalysisMatrix, 
                                    members: List[Member]) -> None:
        """Add referral summary columns."""
        try:
            # Headers
            total_given_col = len(members) + 2
            unique_given_col = len(members) + 3
            
            worksheet.cell(row=1, column=total_given_col, 
                         value=MatrixHeaders.TOTAL_REFERRALS_GIVEN).font = Font(bold=True)
            worksheet.cell(row=1, column=unique_given_col, 
                         value=f"{MatrixHeaders.UNIQUE_REFERRALS_GIVEN} (Total Members = {len(members)})").font = Font(bold=True)
            
            # Data
            for row, member in enumerate(members, start=2):
                stats = matrix.member_statistics.get(member)
                if stats:
                    worksheet.cell(row=row, column=total_given_col, 
                                 value=stats.total_referrals_given).font = Font(bold=True)
                    worksheet.cell(row=row, column=unique_given_col, 
                                 value=stats.unique_referrals_given).font = Font(bold=True)
            
        except Exception as e:
            raise ExportError(f"Error adding referral summary columns: {str(e)}")
    
    def _add_referral_summary_rows(self, worksheet, matrix: AnalysisMatrix, 
                                 members: List[Member]) -> None:
        """Add referral summary rows."""
        try:
            total_received_row = len(members) + 2
            unique_received_row = len(members) + 3
            
            worksheet.cell(row=total_received_row, column=1, 
                         value=MatrixHeaders.TOTAL_REFERRALS_RECEIVED).font = Font(bold=True)
            worksheet.cell(row=unique_received_row, column=1, 
                         value=f"{MatrixHeaders.UNIQUE_REFERRALS_RECEIVED} (Total Members = {len(members)})").font = Font(bold=True)
            
            # Data
            for col, member in enumerate(members, start=2):
                stats = matrix.member_statistics.get(member)
                if stats:
                    worksheet.cell(row=total_received_row, column=col, 
                                 value=stats.total_referrals_received).font = Font(bold=True)
                    worksheet.cell(row=unique_received_row, column=col, 
                                 value=stats.unique_referrals_received).font = Font(bold=True)
            
        except Exception as e:
            raise ExportError(f"Error adding referral summary rows: {str(e)}")
    
    def _add_oto_summary_columns(self, worksheet, matrix: AnalysisMatrix, 
                               members: List[Member]) -> None:
        """Add one-to-one summary columns."""
        try:
            total_oto_col = len(members) + 2
            unique_oto_col = len(members) + 3
            
            worksheet.cell(row=1, column=total_oto_col, 
                         value=MatrixHeaders.TOTAL_OTO).font = Font(bold=True)
            worksheet.cell(row=1, column=unique_oto_col, 
                         value=f"{MatrixHeaders.UNIQUE_OTO} (Total Members = {len(members)})").font = Font(bold=True)
            
            # Data
            for row, member in enumerate(members, start=2):
                stats = matrix.member_statistics.get(member)
                if stats:
                    worksheet.cell(row=row, column=total_oto_col, 
                                 value=stats.total_one_to_ones).font = Font(bold=True)
                    worksheet.cell(row=row, column=unique_oto_col, 
                                 value=stats.unique_one_to_ones).font = Font(bold=True)
                    
        except Exception as e:
            raise ExportError(f"Error adding OTO summary columns: {str(e)}")
    
    def _add_combination_summary_columns(self, worksheet, matrix: AnalysisMatrix, 
                                       members: List[Member]) -> None:
        """Add combination summary columns."""
        try:
            neither_col = len(members) + 2
            oto_only_col = len(members) + 3
            referral_only_col = len(members) + 4
            both_col = len(members) + 5
            
            worksheet.cell(row=1, column=neither_col, value=MatrixHeaders.NEITHER).font = Font(bold=True)
            worksheet.cell(row=1, column=oto_only_col, value=MatrixHeaders.OTO_ONLY).font = Font(bold=True)
            worksheet.cell(row=1, column=referral_only_col, value=MatrixHeaders.REFERRAL_ONLY).font = Font(bold=True)
            worksheet.cell(row=1, column=both_col, value=MatrixHeaders.OTO_AND_REFERRAL).font = Font(bold=True)
            
            # Data
            for row, member in enumerate(members, start=2):
                stats = matrix.member_statistics.get(member)
                if stats:
                    worksheet.cell(row=row, column=neither_col, value=stats.neither_count).font = Font(bold=True)
                    worksheet.cell(row=row, column=oto_only_col, value=stats.oto_only_count).font = Font(bold=True)
                    worksheet.cell(row=row, column=referral_only_col, value=stats.referral_only_count).font = Font(bold=True)
                    worksheet.cell(row=row, column=both_col, value=stats.both_count).font = Font(bold=True)
                    
        except Exception as e:
            raise ExportError(f"Error adding combination summary columns: {str(e)}")
    
    def export_analysis_summary(self, report, file_path: Path) -> None:
        """
        Export a summary of the analysis to Excel.
        
        Args:
            report: Analysis report
            file_path: Output file path
        """
        try:
            workbook = self.excel_handler.create_styled_workbook()
            
            # Create summary sheet
            summary_sheet = workbook.active
            summary_sheet.title = "Analysis Summary"
            
            # Write summary data
            self._write_analysis_summary(summary_sheet, report)
            
            # Create member performance sheet
            member_sheet = workbook.create_sheet("Member Performance")
            self._write_member_performance(member_sheet, report)
            
            # Apply styling to both sheets
            for sheet in workbook.worksheets:
                # Basic styling
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                               top=Side(style="thin"), bottom=Side(style="thin"))
            
            # Save workbook
            self.excel_handler.save_workbook(workbook, file_path)
            
        except Exception as e:
            raise ExportError(f"Failed to export analysis summary: {str(e)}")
    
    def _write_analysis_summary(self, worksheet, report) -> None:
        """Write analysis summary to worksheet."""
        try:
            row = 1
            
            # Title
            worksheet.cell(row=row, column=1, value="BNI PALMS Analysis Summary").font = Font(bold=True, size=16)
            row += 2
            
            # Metadata
            metadata = report.metadata
            worksheet.cell(row=row, column=1, value="Chapter Overview").font = Font(bold=True, size=14)
            row += 1
            
            worksheet.cell(row=row, column=1, value="Total Members:")
            worksheet.cell(row=row, column=2, value=metadata.get('total_members', 0))
            row += 1
            
            worksheet.cell(row=row, column=1, value="Total Referrals:")
            worksheet.cell(row=row, column=2, value=metadata.get('total_referrals', 0))
            row += 1
            
            worksheet.cell(row=row, column=1, value="Total One-to-Ones:")
            worksheet.cell(row=row, column=2, value=metadata.get('total_one_to_ones', 0))
            row += 2
            
            # Matrix summaries
            from src.domain.services.matrix_service import MatrixService
            matrix_service = MatrixService()
            
            for matrix_name, matrix in [
                ("Referral Matrix", report.referral_matrix),
                ("One-to-One Matrix", report.one_to_one_matrix),
                ("Combination Matrix", report.combination_matrix)
            ]:
                worksheet.cell(row=row, column=1, value=matrix_name).font = Font(bold=True, size=12)
                row += 1
                
                summary = matrix_service.get_matrix_summary(matrix)
                worksheet.cell(row=row, column=1, value="Active Members:")
                worksheet.cell(row=row, column=2, value=summary['active_members'])
                row += 1
                
                worksheet.cell(row=row, column=1, value="Total Interactions:")
                worksheet.cell(row=row, column=2, value=summary['total_interactions'])
                row += 2
            
        except Exception as e:
            raise ExportError(f"Error writing analysis summary: {str(e)}")
    
    def _write_member_performance(self, worksheet, report) -> None:
        """Write member performance data to worksheet."""
        try:
            # Headers
            headers = ["Member Name", "Referrals Given", "Referrals Received", 
                      "One-to-Ones", "Total Interactions"]
            
            for col, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            # Data
            row = 2
            for member in report.all_members:
                ref_stats = report.referral_matrix.member_statistics.get(member)
                oto_stats = report.one_to_one_matrix.member_statistics.get(member)
                combo_stats = report.combination_matrix.member_statistics.get(member)
                
                if ref_stats and oto_stats and combo_stats:
                    worksheet.cell(row=row, column=1, value=member.full_name)
                    worksheet.cell(row=row, column=2, value=ref_stats.total_referrals_given)
                    worksheet.cell(row=row, column=3, value=ref_stats.total_referrals_received)
                    worksheet.cell(row=row, column=4, value=oto_stats.total_one_to_ones)
                    worksheet.cell(row=row, column=5, value=combo_stats.total_interactions)
                    row += 1
                    
        except Exception as e:
            raise ExportError(f"Error writing member performance: {str(e)}")
    
    def export_tyfcb_data(self, members: List[Member], tyfcbs: List[TYFCB], file_path: Path) -> None:
        """
        Export TYFCB data to Excel file with detailed breakdown.
        
        Args:
            members: List of all members
            tyfcbs: List of all TYFCB entries
            file_path: Output file path
        """
        try:
            # Initialize TYFCB service
            tyfcb_service = TYFCBService()
            tyfcb_summary = tyfcb_service.generate_tyfcb_summary(members, tyfcbs)
            
            workbook = self.excel_handler.create_styled_workbook()
            
            # Create summary sheet
            summary_sheet = workbook.active
            summary_sheet.title = "TYFCB Summary"
            self._write_tyfcb_summary(summary_sheet, tyfcb_summary)
            
            # Create member breakdown sheet
            member_sheet = workbook.create_sheet("TYFCB by Member")
            self._write_tyfcb_member_breakdown(member_sheet, tyfcb_summary)
            
            # Create detailed transactions sheet
            transactions_sheet = workbook.create_sheet("TYFCB Transactions")
            self._write_tyfcb_transactions(transactions_sheet, tyfcbs)
            
            # Apply styling to all sheets
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                               top=Side(style="thin"), bottom=Side(style="thin"))
                self.excel_handler.auto_adjust_column_widths(sheet)
            
            # Save workbook
            self.excel_handler.save_workbook(workbook, file_path)
            
        except Exception as e:
            raise ExportError(f"Failed to export TYFCB data: {str(e)}")
    
    def _write_tyfcb_summary(self, worksheet, tyfcb_summary: TYFCBSummary) -> None:
        """Write TYFCB summary to worksheet."""
        try:
            row = 1
            
            # Title
            worksheet.cell(row=row, column=1, value="TYFCB Summary Report").font = Font(bold=True, size=16)
            row += 2
            
            # Overall statistics
            worksheet.cell(row=row, column=1, value="Chapter Overview").font = Font(bold=True, size=14)
            row += 1
            
            worksheet.cell(row=row, column=1, value="Total TYFCB Amount:")
            worksheet.cell(row=row, column=2, value=f"${tyfcb_summary.total_amount:,.2f}")
            row += 1
            
            worksheet.cell(row=row, column=1, value="Total TYFCB Count:")
            worksheet.cell(row=row, column=2, value=tyfcb_summary.total_count)
            row += 2
            
            # Within vs Outside Chapter breakdown
            worksheet.cell(row=row, column=1, value="Within Chapter Business").font = Font(bold=True, size=12)
            row += 1
            
            worksheet.cell(row=row, column=1, value="Amount:")
            worksheet.cell(row=row, column=2, value=f"${tyfcb_summary.total_amount_within_chapter:,.2f}")
            row += 1
            
            worksheet.cell(row=row, column=1, value="Count:")
            worksheet.cell(row=row, column=2, value=tyfcb_summary.total_count_within_chapter)
            row += 1
            
            worksheet.cell(row=row, column=1, value="Percentage:")
            worksheet.cell(row=row, column=2, value=f"{tyfcb_summary.within_chapter_percentage:.1f}%")
            row += 2
            
            worksheet.cell(row=row, column=1, value="Outside Chapter Business").font = Font(bold=True, size=12)
            row += 1
            
            worksheet.cell(row=row, column=1, value="Amount:")
            worksheet.cell(row=row, column=2, value=f"${tyfcb_summary.total_amount_outside_chapter:,.2f}")
            row += 1
            
            worksheet.cell(row=row, column=1, value="Count:")
            worksheet.cell(row=row, column=2, value=tyfcb_summary.total_count_outside_chapter)
            row += 1
            
            outside_percentage = 100 - tyfcb_summary.within_chapter_percentage
            worksheet.cell(row=row, column=1, value="Percentage:")
            worksheet.cell(row=row, column=2, value=f"{outside_percentage:.1f}%")
            
        except Exception as e:
            raise ExportError(f"Error writing TYFCB summary: {str(e)}")
    
    def _write_tyfcb_member_breakdown(self, worksheet, tyfcb_summary: TYFCBSummary) -> None:
        """Write TYFCB member breakdown to worksheet."""
        try:
            # Headers
            headers = [
                "Member Name", 
                "Given Within Chapter", "Given Outside Chapter", "Total Given",
                "Received Within Chapter", "Received Outside Chapter", "Total Received"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Data
            row = 2
            for member, stats in tyfcb_summary.member_statistics.items():
                if stats.total_given > 0 or stats.total_received > 0:  # Only show members with TYFCB activity
                    worksheet.cell(row=row, column=1, value=member.full_name)
                    worksheet.cell(row=row, column=2, value=f"${stats.total_given_within_chapter:,.2f}")
                    worksheet.cell(row=row, column=3, value=f"${stats.total_given_outside_chapter:,.2f}")
                    worksheet.cell(row=row, column=4, value=f"${stats.total_given:,.2f}").font = Font(bold=True)
                    worksheet.cell(row=row, column=5, value=f"${stats.total_received_within_chapter:,.2f}")
                    worksheet.cell(row=row, column=6, value=f"${stats.total_received_outside_chapter:,.2f}")
                    worksheet.cell(row=row, column=7, value=f"${stats.total_received:,.2f}").font = Font(bold=True)
                    row += 1
                    
        except Exception as e:
            raise ExportError(f"Error writing TYFCB member breakdown: {str(e)}")
    
    def _write_tyfcb_transactions(self, worksheet, tyfcbs: List[TYFCB]) -> None:
        """Write detailed TYFCB transactions to worksheet."""
        try:
            # Headers
            headers = ["From", "To", "Amount", "Within Chapter", "Description"]
            
            for col, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Data
            row = 2
            for tyfcb in sorted(tyfcbs, key=lambda x: x.amount, reverse=True):  # Sort by amount descending
                giver_name = tyfcb.giver.full_name if tyfcb.giver else "Unknown"
                worksheet.cell(row=row, column=1, value=giver_name)
                worksheet.cell(row=row, column=2, value=tyfcb.receiver.full_name)
                worksheet.cell(row=row, column=3, value=f"${tyfcb.amount:,.2f}")
                worksheet.cell(row=row, column=4, value="Yes" if tyfcb.within_chapter else "No")
                worksheet.cell(row=row, column=5, value=tyfcb.description or "")
                row += 1
                    
        except Exception as e:
            raise ExportError(f"Error writing TYFCB transactions: {str(e)}")
    
    def export_comprehensive_member_report(self, report, file_path: Path) -> None:
        """
        Export comprehensive member report with all key metrics for chapter data only.
        
        Args:
            report: Analysis report containing all data
            file_path: Output file path
        """
        try:
            workbook = self.excel_handler.create_styled_workbook()
            worksheet = workbook.active
            worksheet.title = "Comprehensive Member Report"
            
            # Write the comprehensive member data
            self._write_comprehensive_member_data(worksheet, report)
            
            # Apply styling
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                           top=Side(style="thin"), bottom=Side(style="thin"))
            
            # Auto-adjust column widths
            self.excel_handler.auto_adjust_column_widths(worksheet)
            
            # Save workbook
            self.excel_handler.save_workbook(workbook, file_path)
            
        except Exception as e:
            raise ExportError(f"Failed to export comprehensive member report: {str(e)}")
    
    def _write_comprehensive_member_data(self, worksheet, report) -> None:
        """Write comprehensive member data to worksheet."""
        try:
            # Headers
            headers = [
                "Member Name",
                "Referrals Received", 
                "Referrals Given",
                "One to Ones Done",
                "Number of TYFCBs (Within Chapter)",
                "Total Value of TYFCBs (Within Chapter)"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Initialize TYFCB service for calculations
            from src.domain.services.tyfcb_service import TYFCBService
            tyfcb_service = TYFCBService()
            
            # Filter TYFCBs to only within chapter
            within_chapter_tyfcbs = [tyfcb for tyfcb in report.tyfcbs if tyfcb.within_chapter]
            
            # Generate TYFCB summary for within-chapter data only
            tyfcb_summary = tyfcb_service.generate_tyfcb_summary(report.all_members, within_chapter_tyfcbs)
            
            # Data rows
            row = 2
            for member in sorted(report.all_members, key=lambda m: m.full_name):
                # Get referral statistics
                ref_stats = report.referral_matrix.member_statistics.get(member)
                referrals_received = ref_stats.total_referrals_received if ref_stats else 0
                referrals_given = ref_stats.total_referrals_given if ref_stats else 0
                
                # Get one-to-one statistics
                oto_stats = report.one_to_one_matrix.member_statistics.get(member)
                one_to_ones_done = oto_stats.total_one_to_ones if oto_stats else 0
                
                # Get TYFCB statistics (within chapter only)
                tyfcb_member_stats = tyfcb_summary.member_statistics.get(member)
                if tyfcb_member_stats:
                    num_tyfcbs = tyfcb_member_stats.count_received_within_chapter
                    total_tyfcb_value = tyfcb_member_stats.total_received_within_chapter
                else:
                    num_tyfcbs = 0
                    total_tyfcb_value = 0.0
                
                # Write data
                worksheet.cell(row=row, column=1, value=member.full_name)
                worksheet.cell(row=row, column=2, value=referrals_received)
                worksheet.cell(row=row, column=3, value=referrals_given)
                worksheet.cell(row=row, column=4, value=one_to_ones_done)
                worksheet.cell(row=row, column=5, value=num_tyfcbs)
                worksheet.cell(row=row, column=6, value=f"${total_tyfcb_value:,.2f}")
                
                row += 1
            
            # Add summary row
            row += 1
            worksheet.cell(row=row, column=1, value="TOTALS").font = Font(bold=True)
            
            # Calculate totals
            total_referrals_received = sum(
                report.referral_matrix.member_statistics.get(member, type('obj', (object,), {'total_referrals_received': 0})).total_referrals_received 
                for member in report.all_members
            )
            total_referrals_given = sum(
                report.referral_matrix.member_statistics.get(member, type('obj', (object,), {'total_referrals_given': 0})).total_referrals_given 
                for member in report.all_members
            )
            total_one_to_ones = sum(
                report.one_to_one_matrix.member_statistics.get(member, type('obj', (object,), {'total_one_to_ones': 0})).total_one_to_ones 
                for member in report.all_members
            ) // 2  # Divide by 2 since each one-to-one is counted twice
            
            worksheet.cell(row=row, column=2, value=total_referrals_received).font = Font(bold=True)
            worksheet.cell(row=row, column=3, value=total_referrals_given).font = Font(bold=True)
            worksheet.cell(row=row, column=4, value=total_one_to_ones).font = Font(bold=True)
            worksheet.cell(row=row, column=5, value=tyfcb_summary.total_count_within_chapter).font = Font(bold=True)
            worksheet.cell(row=row, column=6, value=f"${tyfcb_summary.total_amount_within_chapter:,.2f}").font = Font(bold=True)
            
        except Exception as e:
            raise ExportError(f"Error writing comprehensive member data: {str(e)}")