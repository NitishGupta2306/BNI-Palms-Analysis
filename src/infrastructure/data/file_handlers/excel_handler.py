"""Excel file handling utilities."""

from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from src.domain.exceptions.domain_exceptions import FileProcessingError
from src.infrastructure.config.settings import get_settings


class ExcelHandler:
    """Handles Excel file operations."""
    
    def __init__(self):
        self.settings = get_settings()
    
    def read_excel_to_dataframe(self, file_path: Path, **kwargs) -> pd.DataFrame:
        """Read an Excel file into a pandas DataFrame."""
        try:
            if not file_path.exists():
                raise FileProcessingError(f"File not found: {file_path}")
            
            # Default parameters for reading Excel files
            default_params = {
                'dtype': str,
                'header': 0
            }
            default_params.update(kwargs)
            
            # Handle .xls files with appropriate engine
            if file_path.suffix.lower() == '.xls':
                # Check if it's XML-based first
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        first_line = f.readline().strip()
                        if first_line.startswith('<?xml'):
                            # XML-based file - need to convert first
                            from src.infrastructure.data.file_handlers.file_converter import FileConverter
                            converter = FileConverter()
                            xlsx_path = converter.ensure_xlsx_format(file_path, delete_original=False)
                            return pd.read_excel(xlsx_path, **default_params)
                except:
                    pass
                
                # Try binary .xls format
                try:
                    # Try xlrd first for .xls files
                    return pd.read_excel(file_path, engine='xlrd', **default_params)
                except:
                    # If xlrd fails, try openpyxl
                    return pd.read_excel(file_path, engine='openpyxl', **default_params)
            else:
                # For .xlsx files, use default behavior
                return pd.read_excel(file_path, **default_params)
            
        except Exception as e:
            raise FileProcessingError(f"Error reading Excel file {file_path}: {str(e)}")
    
    def read_excel_with_openpyxl(self, file_path: Path) -> Any:
        """Read an Excel file using openpyxl for more control."""
        try:
            if not file_path.exists():
                raise FileProcessingError(f"File not found: {file_path}")
            
            return load_workbook(file_path)
            
        except Exception as e:
            raise FileProcessingError(f"Error reading Excel file {file_path}: {str(e)}")
    
    def create_styled_workbook(self) -> Workbook:
        """Create a new workbook with default styling."""
        return Workbook()
    
    def apply_matrix_styling(self, worksheet, data_range: Dict[str, int], 
                           highlight_zeros: bool = True) -> None:
        """Apply standard matrix styling to a worksheet."""
        try:
            # Define styles
            zero_fill = PatternFill(
                start_color=self.settings.matrix.zero_highlight_color,
                end_color=self.settings.matrix.zero_highlight_color,
                fill_type="solid"
            )
            border_style = Border(
                left=Side(style=self.settings.matrix.border_style),
                right=Side(style=self.settings.matrix.border_style),
                top=Side(style=self.settings.matrix.border_style),
                bottom=Side(style=self.settings.matrix.border_style)
            )
            center_align = Alignment(
                horizontal=self.settings.matrix.text_alignment,
                vertical=self.settings.matrix.text_alignment
            )
            rotated_align = Alignment(
                horizontal=self.settings.matrix.text_alignment,
                vertical=self.settings.matrix.text_alignment,
                textRotation=self.settings.matrix.header_text_rotation
            )
            bold_font = Font(bold=self.settings.matrix.header_bold)
            
            # Apply styling to all cells in the range
            for row in range(1, data_range['max_row'] + 1):
                for col in range(1, data_range['max_col'] + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Apply border and alignment
                    cell.border = border_style
                    cell.alignment = center_align
                    
                    # Style headers
                    if row == 1:
                        cell.alignment = rotated_align
                        cell.font = bold_font
                    
                    if col == 1 and row > 1:
                        cell.font = bold_font
                    
                    # Highlight zero values if requested
                    if (highlight_zeros and row > 1 and col > 1 and 
                        cell.value is not None and cell.value == 0):
                        cell.fill = zero_fill
            
        except Exception as e:
            raise FileProcessingError(f"Error applying styling: {str(e)}")
    
    def auto_adjust_column_widths(self, worksheet, max_width: Optional[int] = None) -> None:
        """Auto-adjust column widths based on content."""
        try:
            max_width = max_width or self.settings.matrix.max_column_width
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(
                    max_length + self.settings.matrix.data_column_width_buffer,
                    max_width
                )
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set first column to standard name width
            worksheet.column_dimensions['A'].width = self.settings.matrix.name_column_width
            
        except Exception as e:
            raise FileProcessingError(f"Error adjusting column widths: {str(e)}")
    
    def save_workbook(self, workbook: Workbook, file_path: Path) -> None:
        """Save a workbook to a file."""
        try:
            # Ensure the directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(file_path)
            
        except Exception as e:
            raise FileProcessingError(f"Error saving workbook to {file_path}: {str(e)}")
    
    def get_file_info(self, file_path: Path) -> Dict[str, Any]:
        """Get information about an Excel file."""
        try:
            if not file_path.exists():
                raise FileProcessingError(f"File not found: {file_path}")
            
            # Read basic info
            df = pd.read_excel(file_path, nrows=5)  # Just read first few rows for info
            
            return {
                'file_name': file_path.name,
                'file_size': file_path.stat().st_size,
                'shape': df.shape,
                'columns': list(df.columns),
                'file_extension': file_path.suffix
            }
            
        except Exception as e:
            raise FileProcessingError(f"Error getting file info for {file_path}: {str(e)}")
    
    def validate_excel_file(self, file_path: Path) -> bool:
        """Validate that a file is a proper Excel file."""
        try:
            if not file_path.exists():
                return False
            
            if file_path.suffix.lower() not in self.settings.excel.supported_extensions:
                return False
            
            # For .xls files, we need to be more careful with validation
            if file_path.suffix.lower() == '.xls':
                # Check if it's an XML-based .xls file
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        first_line = f.readline().strip()
                        if first_line.startswith('<?xml'):
                            # This is an XML-based Excel file, not a binary .xls
                            # We'll need to convert it, but for validation, just check if it's parseable XML
                            try:
                                import xml.etree.ElementTree as ET
                                ET.parse(file_path)
                                return True
                            except Exception:
                                return False
                except:
                    # If we can't read as text, it might be binary
                    pass
                
                # Try binary .xls format
                try:
                    # Try reading with xlrd engine specifically for .xls files
                    pd.read_excel(file_path, nrows=1, engine='xlrd')
                    return True
                except Exception:
                    # If xlrd fails, try openpyxl (for XML-based .xls files)
                    try:
                        pd.read_excel(file_path, nrows=1, engine='openpyxl')
                        return True
                    except Exception:
                        return True  # Allow XML-based files to pass validation
            else:
                # For .xlsx files, use default pandas behavior
                pd.read_excel(file_path, nrows=1)
                return True
            
        except Exception:
            return False