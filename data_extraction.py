from member_extraction import extract_names_from_excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

import pandas as pd

'''GLOBAL VARIABLES:'''
# Styling variables for all cells.
zero_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
border_style = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
center_align = Alignment(horizontal="center", vertical="center")

# Persistent referral matrix (dictionary of dictionaries)
member_names = extract_names_from_excel("Member Names/united_members.xlsx")
referral_matrix = {giver: {receiver: 0 for receiver in member_names} for giver in member_names}
OTO_matrix = {giver: {receiver: 0 for receiver in member_names} for giver in member_names}
combination_matrix = {giver: {receiver: 0 for receiver in member_names} for giver in member_names}


'''CELL STYLING FUNCTIONS:'''
# Function to do complete cell styles for the excel.
def cell_styling(self, bold = False):
    if bold:
        self.font = Font(bold=True)
    self.alignment = center_align
    self.border = border_style

# Function to auto-adjust column widths
def cell_autosizing(ws):
    members = list(referral_matrix.keys())
    for col in range(1, len(members) + 4): # +3 for the title colums.
        max_length = 0
        col_letter = get_column_letter(col)
        # Find the maximum length of content in the column
        for row in range(1, len(members) + 4):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        # Set the column width to fit content
        ws.column_dimensions[col_letter].width = max_length

'''UPDATING MATRIX FUNCTIONS:'''
# Function to process Excel data and update the referral matrix.
def process_referral_excel_data(file_path):
    book = load_workbook(file_path)
    sheet = book.active

    column_slip_type = sheet["C"]
    column_giver_name = sheet["A"]
    column_reciever_name = sheet["B"]

    for cell_slip_type, cell_giver_name, cell_reciever_name in zip(column_slip_type, column_giver_name, column_reciever_name):
        if (cell_reciever_name.value in member_names and
            cell_giver_name.value in member_names and
            cell_slip_type.value == "Referral"):

            giver = cell_giver_name.value
            receiver = cell_reciever_name.value
            referral_matrix[giver][receiver] += 1  # Increment count instead of resetting

    return referral_matrix  # Return updated matrix

# Function to process Excel data and update the OTO matrix.
def process_OTO_excel_data(file_path):
    book = load_workbook(file_path)
    sheet = book.active

    column_slip_type = sheet["C"]
    column_giver_name = sheet["A"]
    column_reciever_name = sheet["B"]

    for cell_slip_type, cell_giver_name, cell_reciever_name in zip(column_slip_type, column_giver_name, column_reciever_name):
        if (cell_reciever_name.value in member_names and
            cell_giver_name.value in member_names and
            cell_slip_type.value == "One to One"):

            giver = cell_giver_name.value
            receiver = cell_reciever_name.value
            OTO_matrix[giver][receiver] += 1  # Increment count instead of resetting

    return OTO_matrix  # Return updated matrix

# Function to update a combination matrix of Referal to OTO
def process_combination_excel_data():
    members = list(referral_matrix.keys())
    for row, giver in enumerate(members, start=2):
        for col, receiver in enumerate(members, start=2):
            # OTO and Referral:
            if ((referral_matrix[giver][receiver] > 0) and (OTO_matrix[giver][receiver] > 0)):
                combination_matrix[giver][receiver] = 3
            # Referral only:
            elif ((referral_matrix[giver][receiver] > 0) and (OTO_matrix[giver][receiver] == 0)):
                combination_matrix[giver][receiver] = 2
            # OTO only:
            elif ((referral_matrix[giver][receiver] == 0) and (OTO_matrix[giver][receiver] > 0)):
                combination_matrix[giver][receiver] = 1
            else:
                combination_matrix[giver][receiver] = 0
    
    return combination_matrix

'''CALCULATION AND EXPORTING FUNCTIONS'''
def export_matrix_to_excel(ws, matrix):
    members = list(matrix.keys())

    # Write headers (X-axis: Receivers)
    ws.cell(row=1, column=1, value="Giver \ Receiver").font = Font(bold=True)
    for col, member in enumerate(members, start=2):
        ws.cell(row=1, column=col, value=member).font = Font(bold=True)

    # Write matrix to excel:
    for row, giver in enumerate(members, start=2):
        ws.cell(row=row, column=1, value=giver).font = Font(bold=True)

        for col, receiver in enumerate(members, start=2):
            value_row = matrix[giver][receiver]
            cell = ws.cell(row=row, column=col, value=value_row)

            if value_row == 0:
                cell.fill = zero_fill  # Highlight zero values
            cell_styling(cell)

# Function to export the matrix to an Excel file
def final_referral_data_to_excel(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Referral Matrix"
    members = list(referral_matrix.keys())

    # Adding Matrix data to workbook
    export_matrix_to_excel(ws, referral_matrix)

    # Adding columns for "Total Referrals Given" and "Unique Referrals Given."
    for row, giver in enumerate(members, start=2):
        ws.cell(row=row, column=1, value=giver).font = Font(bold=True)

        row_values = []
        unique_referrals = 0
        for col, receiver in enumerate(members, start=2):
            value_row = referral_matrix[giver][receiver]

            if value_row > 0:
                unique_referrals += 1
            
            row_values.append(value_row)

        # Creating the new cells.       
        ws.cell(row=1, column=len(members) + 2, value="Total Referrals Given: ").font = Font(bold=True)
        ws.cell(row=1, column=len(members) + 3, value= f"Unique Referals Given: (Total Members = {len(members)})").font = Font(bold=True)

        # Adding calculated values to the new cells.
        avg_cell_given = ws.cell(row=row, column=len(members) + 2, value = sum(row_values))
        unique_cell_given = ws.cell(row = row, column = len(members) + 3, value = unique_referrals)

        # Styling all new cells:
        cell_styling(avg_cell_given, True)
        cell_styling(unique_cell_given, True)

    # Adding rows for "Total Referrals Recieved" and "Unique Referrals Recieved."
    for col, reciever in enumerate(members, start=2):
        col_values = []
        unique_referrals = 0
        for row, giver in enumerate(members, start=2):
            value_col = referral_matrix[giver][reciever]
            if value_col > 0:
                unique_referrals += 1
            col_values.append(value_col)
        
        # Creating the new cells.
        ws.cell(row=len(members) + 2, column=1, value="Total Referrals Recieved: ").font = Font(bold=True)
        ws.cell(row=len(members) + 3, column=1, value= f"Unique Referals Recieved: (Total Members = {len(members)})").font = Font(bold=True)
        
        # Adding calculated values to the new cells:
        avg_cell_recieved = ws.cell(row=len(members) + 2, column=col, value = sum(col_values))
        unique_cell_recieved = ws.cell(row = len(members) + 3, column =col, value = unique_referrals)

        # Styling all new cells:
        cell_styling(avg_cell_recieved, True)
        cell_styling(unique_cell_recieved, True)

    cell_autosizing(ws)
    wb.save(output_file)
    print(f"Referral matrix exported successfully to {output_file}")

def final_OTO_data_to_excel(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "One to One Matrix"
    members = list(OTO_matrix.keys())

    # Adding Matrix data to workbook
    export_matrix_to_excel(ws, OTO_matrix)

    # Adding columns for "Total OTO" and "Unique OTO."
    for row, giver in enumerate(members, start=2):
        ws.cell(row=row, column=1, value=giver).font = Font(bold=True)

        row_values = []
        unique_values = 0
        for col, receiver in enumerate(members, start=2):
            value = OTO_matrix[giver][receiver]
            if value > 0:
                unique_values += 1
            
            row_values.append(value)
        
        # Creating the new cells.
        ws.cell(row = 1, column = len(members) + 2, value = "Total OTO: ").font = Font(bold=True)
        ws.cell(row=1, column=len(members) + 3, value= f"Unique OTO: (Total Members = {len(members)})").font = Font(bold=True)

        # Adding calculated values to the new cells:
        avg_cell = ws.cell(row=row, column=len(members) + 2, value = sum(row_values))
        unique_cell = ws.cell(row = row, column = len(members) + 3, value = unique_values)

        # Styling all new cells:
        cell_styling(avg_cell, True)
        cell_styling(unique_cell, True)

    cell_autosizing(ws)
    wb.save(output_file)
    print(f"OTO matrix exported successfully to {output_file}")

def final_combination_data_to_excel(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Combination Matrix"
    members = list(OTO_matrix.keys())

    # Adding Matrix data to workbook
    process_combination_excel_data()
    export_matrix_to_excel(ws, combination_matrix)

    # Adding columns for "Total OTO" and "Unique OTO."
    for row, giver in enumerate(members, start=2):
        ws.cell(row=row, column=1, value=giver).font = Font(bold=True)

        type_3 = 0
        type_2 = 0
        type_1 = 0
        type_0 = 0

        for col, receiver in enumerate(members, start=2):
            value = combination_matrix[giver][receiver]
            if value == 0:
                type_0 += 1
            elif value == 1:
                type_1 += 1
            elif value == 2:
                type_2 += 1
            else:
                type_3 += 1
        
        # Creating the new cells.
        ws.cell(row=1, column=len(members) + 2, value= "Neither:").font = Font(bold=True)
        ws.cell(row=1, column=len(members) + 3, value= "OTO only:").font = Font(bold=True)
        ws.cell(row=1, column=len(members) + 4, value= "Referral only:").font = Font(bold=True)
        ws.cell(row=1, column=len(members) + 5, value= "OTO and Referral:").font = Font(bold=True)


        # Adding calculated values to the new cells:
        cell_type_0 = ws.cell(row=row, column=len(members) + 2, value = type_0)
        cell_type_1 = ws.cell(row=row, column=len(members) + 3, value = type_1)
        cell_type_2 = ws.cell(row=row, column=len(members) + 4, value = type_2)
        cell_type_3 = ws.cell(row=row, column=len(members) + 5, value = type_3)

        # Styling all new cells:
        cell_styling(cell_type_0, True)
        cell_styling(cell_type_1, True)
        cell_styling(cell_type_2, True)
        cell_styling(cell_type_3, True)

    cell_autosizing(ws)
    wb.save(output_file)
    print(f"Referral matrix exported successfully to {output_file}")

'''CALLING FUNCTION FOR MAIN'''
# Wrapper function to process referrals and export results
def data_extraction(referral_file_path):
    process_referral_excel_data(referral_file_path)
    process_OTO_excel_data(referral_file_path)
    


