import streamlit as st
import pandas as pd
import os
from pathlib import Path
from io import BytesIO
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

def find_header_locations(df):
    """
    Find the locations of the 4 required headers in the dataframe
    Returns a dictionary with header names as keys and (row, col) tuples as values
    """
    headers_to_find = ["Neither:", "OTO only:", "Referral only:", "OTO and Referral:"]
    header_locations = {}

    # Search through the entire dataframe for the headers
    for row_idx in range(len(df)):
        for col_idx in range(len(df.columns)):
            cell_value = df.iloc[row_idx, col_idx]
            if pd.notna(cell_value) and str(cell_value) in headers_to_find:
                header_locations[str(cell_value)] = (row_idx, col_idx)

    # Return only if all 4 headers are found
    if len(header_locations) == 4:
        return header_locations
    else:
        return None

def add_current_referral_column(df, header_locations):
    """
    Add a 'Current Referral' column after the 'OTO and Referral' column
    This column sums 'Referral only' + 'OTO and Referral' values
    """
    df_copy = df.copy()

    # Get the positions of the required columns
    referral_only_row, referral_only_col = header_locations["Referral only:"]
    oto_and_referral_row, oto_and_referral_col = header_locations["OTO and Referral:"]

    # Find the position to insert the new column (after "OTO and Referral")
    new_col_position = oto_and_referral_col + 1

    # Ensure dataframe has enough columns
    while new_col_position >= len(df_copy.columns):
        df_copy[len(df_copy.columns)] = None

    # Add header for the new column
    df_copy.iloc[referral_only_row, new_col_position] = "Current Referral:"

    # Calculate values for each row (skip the header row)
    for row_idx in range(referral_only_row + 1, len(df_copy)):
        referral_only_value = df_copy.iloc[row_idx, referral_only_col]
        oto_and_referral_value = df_copy.iloc[row_idx, oto_and_referral_col]

        # Handle NaN values by treating them as 0
        if pd.isna(referral_only_value):
            referral_only_value = 0
        if pd.isna(oto_and_referral_value):
            oto_and_referral_value = 0

        # Calculate the sum
        current_referral_value = referral_only_value + oto_and_referral_value
        df_copy.iloc[row_idx, new_col_position] = current_referral_value

    return df_copy

def add_last_referral_column(new_matrix_df, old_matrix_df, new_headers, old_headers):
    """
    Add a 'Last Referral' column to the new matrix that looks up
    'Current Referral' values from the old matrix for the same person
    """
    new_df_copy = new_matrix_df.copy()

    # Get positions for new matrix
    new_oto_referral_row, new_oto_referral_col = new_headers["OTO and Referral:"]
    new_current_referral_col = new_oto_referral_col + 1
    new_last_referral_col = new_current_referral_col + 1

    # Get positions for old matrix
    old_oto_referral_row, old_oto_referral_col = old_headers["OTO and Referral:"]
    old_current_referral_col = old_oto_referral_col + 1

    # Ensure dataframe has enough columns
    while new_last_referral_col >= len(new_df_copy.columns):
        new_df_copy[len(new_df_copy.columns)] = None

    # Add header for the new "Last Referral" column
    new_df_copy.iloc[new_oto_referral_row, new_last_referral_col] = "Last Referral:"

    # Create a mapping of person names to their Current Referral values in old matrix
    old_referral_lookup = {}
    for row_idx in range(old_oto_referral_row + 1, len(old_matrix_df)):
        # Get person name from first column (assuming names are in column 0)
        person_name = old_matrix_df.iloc[row_idx, 0]
        if pd.notna(person_name):
            # Get their Current Referral value from old matrix
            current_referral_value = old_matrix_df.iloc[row_idx, old_current_referral_col]
            if pd.isna(current_referral_value):
                current_referral_value = 0
            old_referral_lookup[str(person_name).strip().lower()] = current_referral_value

    # Fill the Last Referral column in new matrix
    for row_idx in range(new_oto_referral_row + 1, len(new_df_copy)):
        # Get person name from new matrix
        person_name = new_df_copy.iloc[row_idx, 0]
        if pd.notna(person_name):
            # Look up their old referral value
            normalized_name = str(person_name).strip().lower()
            last_referral_value = old_referral_lookup.get(normalized_name, 0)
            new_df_copy.iloc[row_idx, new_last_referral_col] = last_referral_value
        else:
            new_df_copy.iloc[row_idx, new_last_referral_col] = 0

    return new_df_copy

def add_change_in_referrals_column(df, headers):
    """
    Add a 'Change in Referrals' column that calculates Current Referral - Last Referral
    Also includes growth/fall indicators
    """
    df_copy = df.copy()

    # Get positions
    oto_referral_row, oto_referral_col = headers["OTO and Referral:"]
    current_referral_col = oto_referral_col + 1
    last_referral_col = current_referral_col + 1
    change_referral_col = last_referral_col + 1

    # Ensure dataframe has enough columns
    while change_referral_col >= len(df_copy.columns):
        df_copy[len(df_copy.columns)] = None

    # Add header for the new column
    df_copy.iloc[oto_referral_row, change_referral_col] = "Change in Referrals:"

    # Calculate change for each row (skip the header row)
    for row_idx in range(oto_referral_row + 1, len(df_copy)):
        current_value = df_copy.iloc[row_idx, current_referral_col]
        last_value = df_copy.iloc[row_idx, last_referral_col]

        # Handle NaN values
        if pd.isna(current_value):
            current_value = 0
        if pd.isna(last_value):
            last_value = 0

        # Calculate the change
        change = current_value - last_value

        # Add indicator for growth/fall
        if change > 0:
            change_text = f"+{change} ↗️"  # Growth indicator
        elif change < 0:
            change_text = f"{change} ↘️"   # Fall indicator
        else:
            change_text = f"{change} ➡️"   # No change indicator

        df_copy.iloc[row_idx, change_referral_col] = change_text

    return df_copy

def add_last_neither_column(new_matrix_df, old_matrix_df, new_headers, old_headers):
    """
    Add a 'Last Neither' column to the new matrix that looks up
    'Neither' values from the old matrix for the same person
    """
    new_df_copy = new_matrix_df.copy()

    # Get positions for new matrix
    new_oto_referral_row, new_oto_referral_col = new_headers["OTO and Referral:"]
    new_current_referral_col = new_oto_referral_col + 1
    new_last_referral_col = new_current_referral_col + 1
    new_change_referral_col = new_last_referral_col + 1
    new_last_neither_col = new_change_referral_col + 1

    # Get positions for old matrix
    old_neither_row, old_neither_col = old_headers["Neither:"]

    # Ensure dataframe has enough columns
    while new_last_neither_col >= len(new_df_copy.columns):
        new_df_copy[len(new_df_copy.columns)] = None

    # Add header for the new "Last Neither" column
    new_df_copy.iloc[new_oto_referral_row, new_last_neither_col] = "Last Neither:"

    # Create a mapping of person names to their Neither values in old matrix
    old_neither_lookup = {}
    for row_idx in range(old_neither_row + 1, len(old_matrix_df)):
        # Get person name from first column (assuming names are in column 0)
        person_name = old_matrix_df.iloc[row_idx, 0]
        if pd.notna(person_name):
            # Get their Neither value from old matrix
            neither_value = old_matrix_df.iloc[row_idx, old_neither_col]
            if pd.isna(neither_value):
                neither_value = 0
            old_neither_lookup[str(person_name).strip().lower()] = neither_value

    # Fill the Last Neither column in new matrix
    for row_idx in range(new_oto_referral_row + 1, len(new_df_copy)):
        # Get person name from new matrix
        person_name = new_df_copy.iloc[row_idx, 0]
        if pd.notna(person_name):
            # Look up their old neither value
            normalized_name = str(person_name).strip().lower()
            last_neither_value = old_neither_lookup.get(normalized_name, 0)
            new_df_copy.iloc[row_idx, new_last_neither_col] = last_neither_value
        else:
            new_df_copy.iloc[row_idx, new_last_neither_col] = 0

    return new_df_copy

def add_change_in_neither_column(df, headers):
    """
    Add a 'Change in Neither' column that calculates Current Neither - Last Neither
    Also includes growth/fall indicators
    """
    df_copy = df.copy()

    # Get positions
    oto_referral_row, oto_referral_col = headers["OTO and Referral:"]
    neither_row, neither_col = headers["Neither:"]

    current_referral_col = oto_referral_col + 1
    last_referral_col = current_referral_col + 1
    change_referral_col = last_referral_col + 1
    last_neither_col = change_referral_col + 1
    change_neither_col = last_neither_col + 1

    # Ensure dataframe has enough columns
    while change_neither_col >= len(df_copy.columns):
        df_copy[len(df_copy.columns)] = None

    # Add header for the new column
    df_copy.iloc[oto_referral_row, change_neither_col] = "Change in Neither:"

    # Calculate change for each row (skip the header row)
    for row_idx in range(oto_referral_row + 1, len(df_copy)):
        current_neither_value = df_copy.iloc[row_idx, neither_col]
        last_neither_value = df_copy.iloc[row_idx, last_neither_col]

        # Handle NaN values
        if pd.isna(current_neither_value):
            current_neither_value = 0
        if pd.isna(last_neither_value):
            last_neither_value = 0

        # Calculate the change
        change = current_neither_value - last_neither_value

        # Add indicator for growth/fall
        # Note: For "Neither", an increase is typically bad (more people with no activity)
        # and a decrease is good (fewer people with no activity)
        if change > 0:
            change_text = f"+{change} ↗️"  # Increase in "neither" (potentially concerning)
        elif change < 0:
            change_text = f"{change} ↘️"   # Decrease in "neither" (potentially good)
        else:
            change_text = f"{change} ➡️"   # No change

        df_copy.iloc[row_idx, change_neither_col] = change_text

    return df_copy

if "uploader_key_new_matrix" not in st.session_state:
    st.session_state.uploader_key_new_matrix = 0
if "uploader_key_old_matrix" not in st.session_state:
    st.session_state.uploader_key_old_matrix = 0

# Directory to save uploaded files
SAVE_DIR_NEW_MATRIX = "New Matrix"
SAVE_DIR_OLD_MATRIX = "Old Matrix"

# Create directories if they don't exist
os.makedirs(SAVE_DIR_NEW_MATRIX, exist_ok=True)
os.makedirs(SAVE_DIR_OLD_MATRIX, exist_ok=True)

st.title("Combination Matrix Comparison")

st.write("## Instructions:")
st.write("""
Note that files must be in .xls or .xlsx format

Follow these steps:
1. Upload the new combination matrix.
2. Upload the old combination matrix.
3. Both matrices will be loaded as dataframes for analysis.""")

st.write("## Upload New Matrix:")
uploaded_new_matrix = st.file_uploader(
    "Choose new combination matrix file",
    type=["xls", "xlsx"],
    key=f"new_matrix_uploader_{st.session_state.uploader_key_new_matrix}"
)

if uploaded_new_matrix:
    file_path = Path(SAVE_DIR_NEW_MATRIX) / uploaded_new_matrix.name
    with open(file_path, "wb") as f:
        f.write(uploaded_new_matrix.getbuffer())

    # Load as dataframe
    try:
        st.session_state.new_matrix_df = pd.read_excel(file_path, header=None)
        st.success(f"New matrix loaded successfully! Shape: {st.session_state.new_matrix_df.shape}")

        # Find header locations
        st.session_state.new_matrix_headers = find_header_locations(st.session_state.new_matrix_df)
        if st.session_state.new_matrix_headers:
            st.success("Found all 4 required headers in new matrix!")
            for header, location in st.session_state.new_matrix_headers.items():
                st.write(f"- {header}: Row {location[0]}, Column {location[1]}")

            # Add Current Referral column
            st.session_state.new_matrix_df = add_current_referral_column(
                st.session_state.new_matrix_df,
                st.session_state.new_matrix_headers
            )
            st.success("Added 'Current Referral' column to new matrix!")
        else:
            st.warning("Could not find all required headers in new matrix")
            st.write("**Debug: Searching for these exact headers:**")
            st.write("- 'Neither:'")
            st.write("- 'OTO only:'")
            st.write("- 'Referral only:'")
            st.write("- 'OTO and Referral:'")
            st.write("**Headers found in your file:**")
            st.write("**Column names:**")
            for i, col in enumerate(st.session_state.new_matrix_df.columns):
                st.write(f"- Column {i}: '{col}'")

            st.write("**First few rows of data:**")
            st.dataframe(st.session_state.new_matrix_df.head(3))

            st.write("**All cells containing ':':**")
            all_headers = []
            for row_idx in range(min(5, len(st.session_state.new_matrix_df))):  # Check first 5 rows
                for col_idx in range(len(st.session_state.new_matrix_df.columns)):
                    cell_value = st.session_state.new_matrix_df.iloc[row_idx, col_idx]
                    if pd.notna(cell_value) and ':' in str(cell_value):
                        all_headers.append(f"'{str(cell_value)}' at Row {row_idx}, Col {col_idx}")
            if all_headers:
                for header in all_headers[:15]:  # Show first 15 cells with colons
                    st.write(f"- {header}")
            else:
                st.write("- No cells containing ':' found in first 5 rows")
    except Exception as e:
        st.error(f"Error loading new matrix: {e}")

st.write("## Upload Old Matrix:")
uploaded_old_matrix = st.file_uploader(
    "Choose old combination matrix file",
    type=["xls", "xlsx"],
    key=f"old_matrix_uploader_{st.session_state.uploader_key_old_matrix}"
)

if uploaded_old_matrix:
    file_path = Path(SAVE_DIR_OLD_MATRIX) / uploaded_old_matrix.name
    with open(file_path, "wb") as f:
        f.write(uploaded_old_matrix.getbuffer())

    # Load as dataframe
    try:
        st.session_state.old_matrix_df = pd.read_excel(file_path, header=None)
        st.success(f"Old matrix loaded successfully! Shape: {st.session_state.old_matrix_df.shape}")

        # Find header locations
        st.session_state.old_matrix_headers = find_header_locations(st.session_state.old_matrix_df)
        if st.session_state.old_matrix_headers:
            st.success("Found all 4 required headers in old matrix!")
            for header, location in st.session_state.old_matrix_headers.items():
                st.write(f"- {header}: Row {location[0]}, Column {location[1]}")

            # Add Current Referral column
            st.session_state.old_matrix_df = add_current_referral_column(
                st.session_state.old_matrix_df,
                st.session_state.old_matrix_headers
            )
            st.success("Added 'Current Referral' column to old matrix!")
        else:
            st.warning("Could not find all required headers in old matrix")
            st.write("**Debug: Searching for these exact headers:**")
            st.write("- 'Neither:'")
            st.write("- 'OTO only:'")
            st.write("- 'Referral only:'")
            st.write("- 'OTO and Referral:'")
            st.write("**Headers found in your file:**")
            all_headers = []
            for row_idx in range(len(st.session_state.old_matrix_df)):
                for col_idx in range(len(st.session_state.old_matrix_df.columns)):
                    cell_value = st.session_state.old_matrix_df.iloc[row_idx, col_idx]
                    if pd.notna(cell_value) and str(cell_value).strip().endswith(':'):
                        all_headers.append(f"'{str(cell_value)}' at Row {row_idx}, Col {col_idx}")
            for header in all_headers[:10]:  # Show first 10 headers found
                st.write(f"- {header}")
    except Exception as e:
        st.error(f"Error loading old matrix: {e}")

# Show status of loaded matrices and add Last Referral column when both are ready
if hasattr(st.session_state, 'new_matrix_df') and hasattr(st.session_state, 'old_matrix_df'):
    if hasattr(st.session_state, 'new_matrix_headers') and hasattr(st.session_state, 'old_matrix_headers'):
        if st.session_state.new_matrix_headers and st.session_state.old_matrix_headers:
            # Add Last Referral column to new matrix if not already added
            if "last_referral_added" not in st.session_state:
                st.session_state.new_matrix_df = add_last_referral_column(
                    st.session_state.new_matrix_df,
                    st.session_state.old_matrix_df,
                    st.session_state.new_matrix_headers,
                    st.session_state.old_matrix_headers
                )
                st.session_state.last_referral_added = True
                st.success("Added 'Last Referral' column to new matrix with values from old matrix!")

            # Add Change in Referrals column if not already added
            if "change_referrals_added" not in st.session_state:
                st.session_state.new_matrix_df = add_change_in_referrals_column(
                    st.session_state.new_matrix_df,
                    st.session_state.new_matrix_headers
                )
                st.session_state.change_referrals_added = True
                st.success("Added 'Change in Referrals' column with growth/fall indicators!")

            # Add Last Neither column if not already added
            if "last_neither_added" not in st.session_state:
                st.session_state.new_matrix_df = add_last_neither_column(
                    st.session_state.new_matrix_df,
                    st.session_state.old_matrix_df,
                    st.session_state.new_matrix_headers,
                    st.session_state.old_matrix_headers
                )
                st.session_state.last_neither_added = True
                st.success("Added 'Last Neither' column with values from old matrix!")

            # Add Change in Neither column if not already added
            if "change_neither_added" not in st.session_state:
                st.session_state.new_matrix_df = add_change_in_neither_column(
                    st.session_state.new_matrix_df,
                    st.session_state.new_matrix_headers
                )
                st.session_state.change_neither_added = True
                st.success("Added 'Change in Neither' column with growth/fall indicators!")

            st.success("Both matrices are loaded and ready for analysis!")

            # Add download button for the final comparison matrix
            st.write("## Download Comparison Results")

            # Create Excel file in memory with formatting
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                st.session_state.new_matrix_df.to_excel(writer, index=False, header=False, sheet_name='Combination_Matrix_Comparison')

                # Get the worksheet to apply formatting
                worksheet = writer.sheets['Combination_Matrix_Comparison']

                # Define styling
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                border_style = Border(left=Side(style="thin"), right=Side(style="thin"),
                                    top=Side(style="thin"), bottom=Side(style="thin"))
                center_align = Alignment(horizontal="center", vertical="center")
                rotated_align = Alignment(horizontal="center", vertical="center", textRotation=90)
                bold_font = Font(bold=True)

                # Get positions to exclude the 4 main columns and added columns from highlighting
                headers = st.session_state.new_matrix_headers
                neither_row, neither_col = headers["Neither:"]
                oto_only_row, oto_only_col = headers["OTO only:"]
                referral_only_row, referral_only_col = headers["Referral only:"]
                oto_referral_row, oto_referral_col = headers["OTO and Referral:"]

                # Columns to exclude from yellow highlighting (the 4 main columns + added columns)
                excluded_cols = {neither_col, oto_only_col, referral_only_col, oto_referral_col}
                added_columns_start = oto_referral_col + 1  # Current Referral column and beyond

                # Apply formatting to all cells
                for row in range(1, len(st.session_state.new_matrix_df) + 1):
                    for col in range(1, len(st.session_state.new_matrix_df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)

                        # Apply borders to all cells
                        cell.border = border_style
                        cell.alignment = center_align

                        # Rotate headers in first row
                        if row == 1:
                            cell.alignment = rotated_align
                            cell.font = bold_font

                        # Apply yellow highlighting to zero values in matrix data only
                        # Exclude: the 4 main columns, added columns, headers, and first column (names)
                        if (row > 1 and col > 1 and  # Skip header row and name column
                            col not in excluded_cols and  # Skip the 4 main columns
                            col < added_columns_start):  # Skip added columns
                            try:
                                if cell.value == 0 or cell.value == 0.0:
                                    cell.fill = yellow_fill
                            except (ValueError, TypeError):
                                pass

                # Auto-adjust column widths for better readability
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 15)  # Cap at 15 for readability
                    worksheet.column_dimensions[column].width = adjusted_width

            # Get the Excel data
            excel_data = output.getvalue()

            # Create download button
            st.download_button(
                label="📊 Download Combination Matrix Comparison",
                data=excel_data,
                file_name="combination_matrix_comparison.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Headers not found in one or both matrices")
elif hasattr(st.session_state, 'new_matrix_df'):
    st.info("New matrix loaded. Please upload the old matrix.")
elif hasattr(st.session_state, 'old_matrix_df'):
    st.info("Old matrix loaded. Please upload the new matrix.")

if st.button("Clear Uploaded Files"):
    # Delete files from directories
    for folder in [SAVE_DIR_NEW_MATRIX, SAVE_DIR_OLD_MATRIX]:
        if os.path.exists(folder):
            for file in os.listdir(folder):
                file_path = os.path.join(folder, file)
                if os.path.isfile(file_path):
                    os.remove(file_path)

    # Clear session state
    for key in ['new_matrix_df', 'old_matrix_df', 'new_matrix_headers', 'old_matrix_headers', 'last_referral_added', 'change_referrals_added', 'last_neither_added', 'change_neither_added']:
        if key in st.session_state:
            del st.session_state[key]

    # Increment uploader keys to reset widgets
    st.session_state.uploader_key_new_matrix += 1
    st.session_state.uploader_key_old_matrix += 1

    # Force rerun
    st.rerun()