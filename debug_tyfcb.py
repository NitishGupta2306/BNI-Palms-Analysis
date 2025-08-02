#!/usr/bin/env python3
"""
Debug script to analyze TYFCB data extraction issues.

This script will help identify why TYFCB entries are not being extracted from Excel files.
"""

import sys
from pathlib import Path

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent / "src"))

import pandas as pd
from openpyxl import load_workbook
from src.shared.constants.app_constants import SlipType, ExcelColumns

def analyze_excel_file(file_path: str):
    """
    Analyze an Excel file to debug TYFCB extraction issues.
    
    Args:
        file_path: Path to the Excel file to analyze
    """
    file_path = Path(file_path)
    
    print(f"Analyzing Excel file: {file_path}")
    print("=" * 60)
    
    if not file_path.exists():
        print(f"❌ ERROR: File does not exist: {file_path}")
        return
    
    try:
        # Load with pandas first to get overview
        print("\n📊 PANDAS ANALYSIS:")
        df = pd.read_excel(file_path, dtype=str)
        print(f"   Shape: {df.shape}")
        print(f"   Columns: {list(df.columns)}")
        
        # Check if we have expected column structure
        slip_type_col = ExcelColumns.SLIP_TYPE.value  # Column C (index 2)
        if df.shape[1] <= slip_type_col:
            print(f"❌ ERROR: File only has {df.shape[1]} columns, but we need at least {slip_type_col + 1} for Slip Type")
            return
        
        # Get the slip type column name
        slip_type_column_name = df.columns[slip_type_col]
        print(f"   Slip Type column (index {slip_type_col}): '{slip_type_column_name}'")
        
        # Analyze slip types
        slip_types = df.iloc[:, slip_type_col].dropna()
        unique_slip_types = slip_types.unique()
        
        print(f"\n🔍 SLIP TYPES FOUND:")
        print(f"   Total non-null entries: {len(slip_types)}")
        print(f"   Unique slip types: {len(unique_slip_types)}")
        
        for slip_type in unique_slip_types:
            count = len(slip_types[slip_types == slip_type])
            repr_value = repr(slip_type)  # Shows quotes and escape characters
            print(f"   - {repr_value}: {count} occurrences")
            
            # Check exact match with our constant
            if slip_type == SlipType.TYFCB.value:
                print(f"     ✅ EXACT MATCH with SlipType.TYFCB.value ('{SlipType.TYFCB.value}')")
            else:
                print(f"     ❌ NO MATCH with SlipType.TYFCB.value ('{SlipType.TYFCB.value}')")
                
                # Check for common variations
                if isinstance(slip_type, str):
                    stripped = slip_type.strip()
                    lower = slip_type.lower()
                    upper = slip_type.upper()
                    
                    print(f"     Variations:")
                    print(f"       - Stripped: {repr(stripped)}")
                    print(f"       - Lower: {repr(lower)}")
                    print(f"       - Upper: {repr(upper)}")
                    
                    if stripped == SlipType.TYFCB.value:
                        print(f"       ✅ STRIPPED version matches!")
                    if upper == SlipType.TYFCB.value:
                        print(f"       ✅ UPPER version matches!")
                    if lower == SlipType.TYFCB.value.lower():
                        print(f"       ✅ LOWER version matches!")
        
        print(f"\n📋 EXPECTED VALUES:")
        print(f"   SlipType.REFERRAL.value = {repr(SlipType.REFERRAL.value)}")
        print(f"   SlipType.ONE_TO_ONE.value = {repr(SlipType.ONE_TO_ONE.value)}")
        print(f"   SlipType.TYFCB.value = {repr(SlipType.TYFCB.value)}")
        
        # Now load with openpyxl to see raw cell values
        print(f"\n🔬 OPENPYXL ANALYSIS (Raw Cell Values):")
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        print(f"   Active sheet: '{sheet.title}'")
        print(f"   Max row: {sheet.max_row}")
        print(f"   Max column: {sheet.max_column}")
        
        # Check first 20 rows of slip type column
        slip_type_col_letter = chr(ord('A') + slip_type_col)  # Convert index to letter
        print(f"\n   First 20 rows of column {slip_type_col_letter} (Slip Type):")
        
        tyfcb_found = 0
        for row_idx in range(1, min(21, sheet.max_row + 1)):
            cell = sheet.cell(row=row_idx, column=slip_type_col + 1)  # openpyxl is 1-indexed
            value = cell.value
            
            print(f"     Row {row_idx}: {repr(value)}")
            
            if value == SlipType.TYFCB.value:
                tyfcb_found += 1
                print(f"       ✅ EXACT MATCH!")
            elif isinstance(value, str) and value.strip() == SlipType.TYFCB.value:
                print(f"       ⚠️  MATCHES after stripping whitespace")
            elif isinstance(value, str) and value.upper() == SlipType.TYFCB.value:
                print(f"       ⚠️  MATCHES after converting to uppercase")
        
        print(f"\n📊 SUMMARY:")
        print(f"   TYFCB entries with exact match: {tyfcb_found}")
        
        # Check for potential TYFCB rows by looking at all columns
        print(f"\n🔍 SEARCHING FOR POTENTIAL TYFCB ROWS:")
        potential_tyfcb_rows = []
        for row_idx in range(2, sheet.max_row + 1):  # Skip header
            row_values = []
            for col_idx in range(1, min(8, sheet.max_column + 1)):  # Check first 7 columns
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_values.append(cell.value)
            
            # Check if any cell contains "TYFCB"
            row_text = ' '.join([str(v) if v else '' for v in row_values])
            if 'TYFCB' in row_text.upper():
                potential_tyfcb_rows.append((row_idx, row_values))
        
        if potential_tyfcb_rows:
            print(f"   Found {len(potential_tyfcb_rows)} rows that mention TYFCB:")
            for row_idx, values in potential_tyfcb_rows[:5]:  # Show first 5
                print(f"     Row {row_idx}: {values}")
        else:
            print("   No rows found that mention TYFCB")
        
    except Exception as e:
        print(f"❌ ERROR analyzing file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python debug_tyfcb.py <path_to_excel_file>")
        print("\nExample:")
        print("  python debug_tyfcb.py '/path/to/your/excel/file.xlsx'")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    analyze_excel_file(excel_file)